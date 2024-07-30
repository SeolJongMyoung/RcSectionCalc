import openpyxl as op
from openpyxl import Workbook
from string import ascii_uppercase
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side, Alignment, PatternFill
from rcsec_usd_calc import CalcReinfoeceConcrete

def create_excel_output(calculator):
    wb = Workbook()
    wsout = wb.active
    
    # Set up basic formatting
    alpalist = list(ascii_uppercase)
    for i in range(1, 100):
        wsout.row_dimensions[i].height = 15
    for i in alpalist:
        wsout.column_dimensions[i].width = 3.0
    font_format = Font(size=9, name='굴림체')
    for rows in wsout["A1":"Z100"]:
        for cell in rows:
            cell.font = font_format

    # Section 1: Basic Information
    wsout['B2'].value = '1) 단면제원 및 설계가정'
    wsout['C3'].value = f"fck = {calculator.f_ck} MPa, fy = {calculator.f_y} MPa, Øf = {calculator.pi_f_r:.2f}, Øv = {calculator.pi_v:.2f}, Es = {calculator.E_s} MPa"

    # Section 2: Concrete Material Constants
    for r in range(4,6) :                        # 4~5행 글짜 위치 중앙으로 정렬
        for c in range(3,24) :                      
            wsout.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')  
    for r in range(4,6) :                        # 4~5행 테두리 그리기
        for c in range(3,24) :                      
            wsout.cell(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    for r in range(4,6) :                        # 셀 병합 C4~L4, C5~L5
        for c in [3,6,9,12] :                       
            c1 = c + 2
            wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
    for r in range(4,6) :                         # 셀 병합 O4~T4, O5~T5  
        for c in [15,20] :                          
            c1 = c + 4
            wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)    
    for c in range(3,24) :                         # 셀 채우기 C4~T5
        wsout.cell(4,c).fill = PatternFill(fill_type='solid', fgColor='0FFFF0')
    wsout['C4'].value = 'B(mm)'
    wsout['F4'].value = 'H(mm)'
    wsout['I4'].value = 'd(mm)'
    wsout['L4'].value = '피복(mm)'
    wsout['O4'].value = 'Mu(N.mm)'
    wsout['T4'].value = 'Vu(N)'
    wsout['C5'].value = calculator.beam_b
    wsout['F5'].value = calculator.beam_h
    wsout['I5'].value = f"{calculator.d_eff:.1f}"
    wsout['L5'].value = f"{calculator.d_c:.1f}"
    wsout['O5'].value = calculator.Mu_nm
    wsout['T5'].value = calculator.Vu_n
    wsout['B7'].value = '2) 콘크리트 재료상수'
    wsout['C8'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수'
    wsout['P8'].value = '='
    wsout['Q8'].value = calculator.beta_1

    # Section 3: Strength reduction factor
    wsout['B10'].value = '3) 강도감소계수(Ø) 산정'
    wsout['C11'].value = f"T = As x fy = {calculator.as_use:.3f} x {calculator.f_y:.3f} = {calculator.tension_force:.3f} N"
    wsout['C12'].value = f"C = 0.85 x fck x a x b = 0.85 x {calculator.f_ck:.3f} x a x {calculator.beam_b:.3f}"
    wsout['C13'].value = f"T = C 이므로, a = {calculator.a:.3f} mm, c = {calculator.a:.3f} / β1 = {calculator.a:.3f} / {calculator.beta_1:.3f} = {calculator.c:.3f} mm"
    wsout['C14'].value = f"εy = fy / Es = {calculator.f_y:.2f} / {calculator.E_s:.2f} = {calculator.epsilon_y:.5f}"
    wsout['C15'].value = f"εt = 0.00300 x (dt - c) / c = 0.00300 x ({calculator.d_eff:.3f} - {calculator.c:.2f}) / {calculator.c:.2f} = {calculator.epsilon_t:.5f}"
    if calculator.epsilon_t >= 0.005 :
        wsout['C16'].value = f"εt ≥ 0.0050 이므로 {calculator.epsilon_t_result}이며, Ø = {calculator.pi_f_r:.2f} 를 적용한다"
    else :
        wsout['C16'].value = f"εt < 0.0050 이므로 {calculator.epsilon_t_result}이며, Ø = {calculator.pi_f_r:.2f} 를 적용한다"

    # Section 4: Required Reinforcement Calculation
    wsout['B18'].value = '4) 필요철근량 산정'
    wsout['C19'].value = 'Mu / Øf = As x fy  x (d - a / 2)              ----------------   ①'
    wsout['C20'].value = ' a = As x fy  / ( 0.85 x fck x b)             ----------------   ②'
    wsout['C21'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
    wsout['E22'].value = ' fy²                                Mu'
    wsout['C23'].value = f" ────────── As² - fy x d x As + ───  = 0 ,   Asreq = {calculator.as_req:.3f} mm²"
    wsout['C24'].value = ' 2 x 0.85 x fck x b                         Øf '
    
    # Section 5: Used Reinforcement
    wsout['B26'].value = f"5) 사용철근량 : Asuse = {calculator.as_use:.1f} mm², 철근도심 : dc = {calculator.d_eff:.1f} mm [ 사용율 = {calculator.as_use/calculator.as_req:.3f} ]"
    wsout['F27'].value = f"1단 : {calculator.rebar_id} {calculator.as_dia1} - {calculator.as_num1} EA (= {calculator.as_use1*calculator.as_num1:.1f} mm², dc1 = {calculator.dc_1:.1f} mm)"
    wsout['F28'].value = f"2단 : {calculator.rebar_id} {calculator.as_dia2} - {calculator.as_num2} EA (= {calculator.as_use2*calculator.as_num2:.1f} mm², dc2 = {calculator.dc_2:.1f} mm)"
    wsout['F29'].value = f"3단 : {calculator.rebar_id} {calculator.as_dia3} - {calculator.as_num3} EA (= {calculator.as_use3*calculator.as_num3:.1f} mm², dc3 = {calculator.dc_3:.1f} mm)"

    # Section 6: Reinforcement Check
    wsout['B31'].value = "6) 철근비 검토"
    wsout['C32'].value = f"ρmin : 1.4 / fy          = {calculator.lo_min_1:.6f} "
    wsout['C33'].value = f"       0.25 x √fck / fy  = {calculator.lo_min_2:.6f},  ρmin = {calculator.lo_min:.6f} 적용"
    wsout['C34'].value = f"ρmax = 0.75 x ρb = 0.75 x (0.85 x β1 x fck / fy) x (6,000 / (6,000 + fy)) = {calculator.lo_bal:.6f}" 
    wsout['C35'].value = f"ρuse = As / ( b x d ) = {calculator.lo_use:.6f} " 
    if calculator.lo_use >= calculator.lo_min :
        if calculator.lo_use < calculator.lo_max :
            wsout['C36'].value = f"ρmax ≥ ρuse ≥ ρmin --> 최소철근비, 최대철근비 만족   ∴ O.K"
        else:
            wsout['C36'].value = f"ρmax < ρuse ≥ ρmin --> 최소철근비 만족, 최대 철근비 불만족   ∴ N.G"
    else :
        if calculator.lo_use < calculator.lo_max :
            wsout['C36'].value = f"ρmax ≥ ρuse < ρmin --> 최소철근비 불만족,  최대 철근비 만족   ∴ N.G"
            if calculator.lo_use >= calculator.lo_min_3 :
                wsout['C37'].value = f"ρuse ≥ 4 x ρreq / 3 = {calculator.lo_min_3:.6f} --> 최소철근비 만족   ∴ O.K"
            else :
                wsout['C37'].value = f"ρuse < 4 x ρreq / 3 = {calculator.lo_min_3:.6f} --> 최소철근비 불만족   ∴ N.G"
        else :    
            wsout['C36'].value = f"ρmax < ρuse < ρmin --> 최소철근비, 최대 철근비 불만족   ∴ N.G"

    # Section 7: Neutral Axis Depth Check
    wsout['B44'].value = "7) 중립축 깊이 검토"
    wsout['C45'].value = f"C = a / β1 = {calculator.a:.1f} / {calculator.beta_1:.3f} = {calculator.c:.1f} mm"

    # Section 8: Tensile Reinforcement Strain
    wsout['B51'].value = "8) 인장철근 변형률"
    wsout['C52'].value = f"εt = {calculator.epsilon_t:.5f}"
    wsout['C53'].value = f"εy = {calculator.epsilon_y:.5f}"
    wsout['C54'].value = f"Strain check result: {calculator.epsilon_t_result}"

    # Section 9: Design Flexural Strength Calculation
    wsout['B57'].value = "9) 설계 휨강도 산정"
    wsout['C58'].value = f"Mr = As x Øf x fy x (d - a / 2)"
    wsout['C59'].value = f"   = {calculator.as_use:.1f} x {calculator.pi_f_r:.2f} x {calculator.f_y} x ({calculator.d_eff:.1f} - {calculator.a:.1f} / 2)"
    if calculator.M_r > calculator.Mu_nm:
        wsout['C60'].value = f"   = {calculator.M_r:.1f} N.mm ≥ Mu = {calculator.Mu_nm:.1f} N.mm  ∴ O.K  [S.F = {calculator.M_sf:.3f}]"
    else:
        wsout['C60'].value = f"   = {calculator.M_r:.1f} N.mm < Mu = {calculator.Mu_nm:.1f} N.mm  ∴ N.G  [S.F = {calculator.M_sf:.3f}]"

    # Save the workbook
    wb.save('Calc_As_Output.xlsx')
    wb.close()

def main():
    file_path = "D:\Python\RC_Section_Calculation/Calc_As_input.xlsx"
    calculator = CalcReinfoeceConcrete(file_path)
    calculator.calculate()
    create_excel_output(calculator)
    print('Excel file "Calc_As_Output.xlsx" has been created.')

if __name__ == "__main__":
    main()