#import xlwings as xw
import openpyxl as op
from openpyxl import load_workbook, Workbook
from string import ascii_uppercase
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.styles.colors import Color
from sec_back_ver02 import *

#--------------------------------------------------
#    직사각형 단철근보 단면검토 출력 (도로교설계기준 2010)
#--------------------------------------------------

calc.calmoment()
wb = Workbook()
wsout = wb.active
#wsout = wb.create_sheet("Result")
alpalist = list(ascii_uppercase)
for i in range(1,100) :                     # 헹 크기 지정
    wsout.row_dimensions[i].height = 15
for i in alpalist :                         # 열 크기 지정
    wsout.column_dimensions[i].width = 3.0
font_format = Font(size=9, name = '굴림체')
for rows in wsout["A1":"Z100"] :            # 기본 폰트를 size=9, 굴림체로 변경
    for cell in rows :
        cell.font = font_format
wsout['B2'].value = '1) 단면제원 및 설계가정'
wsout['C3'].value ="fck = %d Mpa, fy = %d Mpa, Øc = %1.2f, Øs = %1.2f, Es = %d Mpa" %(calc.fck, calc.fy, calc.Øc, calc.Øs, calc.Es)
for r in range(4,6) :                        # 4~5행 글짜 위치 중앙으로 정렬
    for c in range(3,24) :                      
        wsout.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')  
for r in range(4,6) :                        # 4~5행 테두리 그리기
    for c in range(3,24) :                      
        wsout.cell(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
for r in range(4,6) :                        # 셀 병합 C4~L4, C5~L5
    for c in [3,6,9,12] :                       
        c1 = c + 2
        wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
for r in range(4,6) :                         # 셀 병합 O4~T4, O5~T5  
    for c in [15,20] :                          
        c1 = c + 4
        wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)    
for c in range(3,24) :                         # 셀 채우기 C4~T5
    wsout.cell(4,c).fill = PatternFill(fill_type='solid', fgColor='0FFFF0')
wsout['C4'].value = 'B(mm)'
wsout['F4'].value = 'H(mm)'
wsout['I4'].value = 'd(mm)'
wsout['L4'].value = '피복(mm)'
wsout['O4'].value = 'Mu(N.mm)'
wsout['T4'].value = 'Vu(N)'
wsout['C5'].value = calc.B
wsout['F5'].value = calc.H
wsout['I5'].value = calc.D
wsout['L5'].value = calc.Dc
wsout['O5'].value = calc.Mun
wsout['T5'].value = calc.Vun
wsout['B7'].value = '2) 콘크리트 재료상수'
for r in range(8,19) :
    wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
wsout['C8'].value = 'n      : 상승 곡선부의 형상을 나타내는 지수'
wsout['P8'].value = '='
wsout['Q8'].value = calc.nε
wsout['C9'].value = 'εco,r : 최대응력에 처음 도달했을때의 변형률'
wsout['P9'].value = '='
wsout['Q9'].value = calc.εco
wsout['C10'].value = 'εcu,r : 극한변형률'
wsout['P10'].value = '='
wsout['Q10'].value = calc.εcu
wsout['C11'].value = 'αcc   : 유효계수'
wsout['P11'].value = '='
wsout['Q11'].value = calc.αcc
wsout['C12'].value = 'fcd   : 콘크리트 설계압축강도'
wsout['P12'].value = '='
wsout['Q12'].value = calc.fcd
wsout['S12'].value = 'MPa'
wsout['C13'].value = 'fcm   : 평균압축강도(fck+Δf)'
wsout['P13'].value = '='
wsout['Q13'].value = calc.fcm
wsout['S13'].value = 'MPa'
wsout['C14'].value = 'Ec    : 콘크리트 탄성계수'
wsout['P14'].value = '='
wsout['Q14'].value = calc.Ec
wsout['S14'].value = 'MPa'
wsout['C15'].value = 'α     : 압축합력의 평균 응력계수'
wsout['P15'].value = '='
wsout['Q15'].value = calc.α
wsout['C16'].value = 'β     : 압축합력의 작용점 위치계수'
wsout['P16'].value = '='
wsout['Q16'].value = calc.β
wsout['C17'].value = 'η     : 등가 사각형 응력 블록의 크기계수'
wsout['P17'].value = '='
wsout['Q17'].value = calc.η
wsout['C18'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수(2β)'
wsout['P18'].value = '='
wsout['Q18'].value = calc.β1

wsout['B20'].value = '3) 철근 재료상수'
for r in range(21,23) :
    wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
wsout['C21'].value = 'fyd    : 설계인장강도 ( Φs fy )'
wsout['P21'].value = '='
wsout['Q21'].value = calc.fyd
wsout['S21'].value = 'MPa'
wsout['C22'].value = 'εyd    : 설계 항복 변형률 ( fyd / Es )'
wsout['P22'].value = '='
wsout['Q22'].value = calc.εyd

wsout['B24'].value = '4) 필요철근량 산정'
wsout['C25'].value = 'Mu = As x fyd  x (d - a / 2)              ----------------   ①'
wsout['C26'].value = ' a = As x fyd  / ( η x fcd x b)          ----------------   ②'
wsout['C27'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
wsout['F28'].value = ' fyd²'
wsout['C29'].value = ' ──────────── As² - fyd x d x As + Mu = 0 ,   Asreq = %5.3fmm²' %(calc.Asreq)
wsout['D30'].value = ' 2 x η x fcd x b'

wsout['B32'].value ="5) 사용철근량 : Asuse =  %6.1f mm², 철근도심 : dc =  %5.1f mm [ 사용율 = %3.3f ]" %(calc.Asuse, calc.Dc, calc.Asuse/calc.Asreq)
wsout['F33'].value = "1단 : %c %d - %d EA (=  %5.1f mm², dc1 =  %5.1f mm)" %(calc.rebarid, calc.AsDia1, calc.AsNum1, calc.Asuse1*calc.AsNum1, calc.Dc1)
wsout['F34'].value = "2단 : %c %d - %d EA (=  %5.1f mm², dc2 =  %5.1f mm)" %(calc.rebarid, calc.AsDia2, calc.AsNum2, calc.Asuse2*calc.AsNum2, calc.Dc2)
wsout['F35'].value = "3단 : %c %d - %d EA (=  %5.1f mm², dc3 =  %5.1f mm)" %(calc.rebarid, calc.AsDia3, calc.AsNum3, calc.Asuse3*calc.AsNum3, calc.Dc3)

wsout['B37'].value ="6) 철근량 검토"
wsout['C38'].value ="As,min = (0.25 √fck / fy) x b x d = %6.1f mm²" %(calc.Asmin1)
wsout['D39'].value ="   = (1.4 / fy) x b x d = %6.1f mm²" %(calc.Asmin2) 
wsout['D40'].value ="   = As_req x 4 / 3 = %6.1f mm²" %(calc.Asmin3) 
if calc.Asuse >= calc.Asmin :
    wsout['C41'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²  ∴ O.K" %(calc.Asuse, calc.Asmin)
else :
    wsout['C41'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²  ∴ N.G" %(calc.Asuse, calc.Asmin)
if calc.Asuse <= calc.Asmax :
    wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² ≥  As,use = %6.1f mm²  ∴ O.K" %(calc.Asmax, calc.Asuse)
else :
    wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² <  As,use = %6.1f mm²  ∴ N.G" %(calc.Asmax, calc.Asuse)

wsout['B44'].value ="7) 중립축 깊이 검토"
wsout['C45'].value ="Cmax = (δ x εcu / 0.0033 - 0.6) x d" 
wsout['D46'].value =" = (%2.1f x %2.5f / 0.0033 - 0.6) x %4.1f = %4.1f mm" %(calc.δ, calc.εcu, calc.D, calc.c_max)
wsout['C47'].value ="C = Φs x As x fy / (α x Φc x 0.85 x fck x b)" 
wsout['C48'].value ="  = %2.2f x %6.1f x %d / (%2.2f x %2.2f x 0.85 x %d x %5.1f)" %(calc.Øs, calc.Asuse, calc.fy, calc.α, calc.Øc, calc.fck, calc.B)
if calc.cc <= calc.c_max :
    wsout['C49'].value ="  = %6.1f mm < Cmax = %6.1f mm  ∴ O.K" %(calc.cc, calc.c_max)
else :
    wsout['C49'].value ="  = %6.1f mm ≥ Cmax = %6.1f mm  ∴ N.G" %(calc.cc, calc.c_max)

wsout['B51'].value ="8) 인장철근 변형률"
wsout['C52'].value ="εs = (d - C) / C x εcu" 
wsout['C53'].value ="    = (%4.1f - %4.1f) / %4.1f x %2.5f = %2.5f " %(calc.D, calc.cc, calc.cc, calc.εcu, calc.εs)
wsout['C54'].value ="εyd = Φs x fy / Es" 
if calc.εyd <= calc.εs :
    wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  ≤ εs  ∴ 항복가정 O.K" %(calc.Øs, calc.fy, calc.Es, calc.εyd)
else :
    wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  > εs  ∴ 항복가정 N.G" %(calc.Øs, calc.fy, calc.Es, calc.εyd)

wsout['B57'].value ="9) 설계 휨강도 산정"
wsout['C58'].value ="Mr = As x Φs x fy x (d - β x c)" 
wsout['C59'].value ="   = %5.1f x %2.2f x %d x (%4.1f - %2.2f x %4.1f)" %(calc.Asuse, calc.Øs, calc.fy, calc.D, calc.β, calc.cc)
if calc.Mr > calc.Mun :
    wsout['C60'].value ="   = %10.1f N.mm  ≥ Mu = %10f.1 N.mm  ∴ O.K  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
else :
    wsout['C60'].value ="   = %10.1f N.mm  < Mu = %10.1f N.mm  ∴ N.G  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)

#------------------------------
#          전단력 검토
#------------------------------
calc.calshear()
wsout['B62'].value ="10) 전단검토"
wsout['C63'].value ="Vcd = [0.85 x Φc x k x (p x fck)⅓ + 0.15 x fn] x b x d" 
wsout['C64'].value ="    = [0.85 x %1.2f x %1.3f x (%1.6f x %3.1f)⅓ + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.k, calc.ρs, calc.fck, calc.fn, calc.B, calc.D)
wsout['C65'].value ="    = %9.0f N" %(calc.Vc) 
wsout['D66'].value ="k = 1 + √(200 / d) = 1 + √(200 / %4.1f) = %1.3f (≤ 2.000)" %(calc.D, calc.k)
wsout['D67'].value ="p = As / (b x d) = %5.1f / (%4.1f x %4.1f) = %1.6f ≤ 0.0200" %(calc.Asuse, calc.B, calc.D, calc.ρ) 
wsout['D68'].value ="fn = Nu / Ac = %4.3f MPa (≤ 0.2 Φc fck = %4.3f MPa, 압축의경우+)" %(calc.fnn, calc.fnmax)
wsout['C69'].value ="Vcd,min = (0.4 x Φc x fctk + 0.15 x fn] x b x d" 
wsout['C70'].value ="        = (0.4 x %1.2f x %3.3f + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.fctk, calc.fn, calc.B, calc.D)
wsout['C71'].value ="        = %9.0f N" %(calc.Vcdmin)

if calc.Vcd >= calc.Vun :
    wsout['C72'].value ="Vcd = %9.0f N ≥ Vu = %9.0f N  ∴ 전단보강 불필요" %(calc.Vcd, calc.Vun)
else :
    wsout['C72'].value ="Vcd = %9.0f N < Vu = %9.0f N  ∴ 전단보강 필요" %(calc.Vcd, calc.Vun)
    wsout['C74'].value ="cotΘ = %1.3f ( Θ = %3.1f˚)" %(calc.cotθ, calc.θ)
    if calc.cotθ < 1 :
        wsout['L74'].value ="1.0 > cotΘ → Θ 불만족  ∴ N.G"
    elif calc.cotθ > 2.5 :    
        wsout['L74'].value ="2.5 ≤ cotΘ → Θ 불만족  ∴ N.G"
    else :
        wsout['L74'].value ="1.0 ≤ cotΘ ≤ 2.5 → Θ 만족  ∴ O.K"
    wsout['C76'].value ="Vsd = (Φs x fvy x Av x z / s) x cotθ"
    wsout['C77'].value ="    = (%1.2f x %4.1f x %4.1f x %4.1f / %4.1f) x %1.3f" %(calc.Øs, calc.fy, calc.Avs, calc.z, calc.AvSpace, calc.cotθ)
    if calc.Vd >= calc.Vun :
        wsout['C78'].value ="    = %d N ≥ Vu = %d N  ∴ O.K" %(calc.Vd, calc.Vun)
    else :
        wsout['C78'].value ="    = %d N < Vu = %d N  ∴ N.G" %(calc.Vd, calc.Vun)    
    wsout['C79'].value ="Vsd,max = (v x Φc x fck x b x z) / (cotθ + tanθ)"
    wsout['C80'].value ="        = (%1.3f x %1.2f x %3.1f x %4.1f x %4.1f) / ( %1.3f + %1.3f )" %(calc.ν, calc.Øc, calc.fck, calc.B, calc.z, calc.cotθ, calc.tanθ)
    if calc.Vd <= calc.Vdmax :
        wsout['C81'].value ="        = %d N ≤ Vsd = %d N  ∴ O.K" %(calc.Vdmax, calc.Vd)
    else :
        wsout['C81'].value ="        = %d N > Vsd = %d N  ∴ N.G" %(calc.Vdmax, calc.Vd)    






#--------------------------------------
#          사용성 검토(균열검토)
#--------------------------------------
#calc.calservice()
wb.save('Calc_As_Output.xlsx')
wb.close()